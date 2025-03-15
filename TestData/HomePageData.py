import os

import openpyxl

class HomePageData:
    # test = [{"n":"Rahul", "f":"Shetty", "g":"Female"}, {"n":"Rahul", "f":"Shetty", "g":"Male"}]

    @classmethod
    def getTestData(cls, test_case_name):
        dic = dict()
        base_path = os.path.dirname(os.path.abspath(__file__))
        # Формируем путь к файлу Excel
        file_path = os.path.join(base_path, "PythonDemo.xlsx")
        book = openpyxl.load_workbook(file_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    print(sheet.cell(row=i, column=j).value)
        return [dic]